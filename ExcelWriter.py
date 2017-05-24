#!python3.exe
# Tested with python 3.6.1

from openpyxl import Workbook
from openpyxl import load_workbook

class ExcelWriter:
    def __init__(self, filename):
        self.__first_run    = True
        self.last_row       = 2
        self.filename       = filename
        self.workbook       = Workbook()
        self.worksheet      = self.workbook.active

    def close(self):
        self.workbook.save(self.filename)

    def save_header(self, object):
        row = 1
        col = 1

        for key in list(object[0][0].keys()): # default parameters
            self.worksheet.cell(column=col, row=row, value=key)
            col += 1

        for key in list(object[0][1].keys()): # extended parameters
            self.worksheet.cell(column=col, row=row, value=key)
            col += 1

    def save(self, object):
        row = self.last_row
        col = 1

        if self.__first_run:
            # Only write the column names one time at the top of the file
            self.save_header(object)
            self.__first_run = False
        else:
            self.workbook = load_workbook(self.filename)
            self.worksheet = self.workbook.active

        for line in object:
            for value in list(line[0].values()):
                self.worksheet.cell(column=col, row=self.last_row, value=value)
                col += 1
            for value in list(line[1].values()):
                self.worksheet.cell(column=col, row=self.last_row, value=value)
                col += 1
            col = 1
            self.last_row += 1

        self.close()
